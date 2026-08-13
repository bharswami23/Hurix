from pathlib import Path
from openpyxl import load_workbook

tol = 1e-6

ref_file = Path("/solution/oracle_results.xlsx")
agent_file = Path("/logs/agent/results.xlsx")

if not agent_file.exists():
    Path("/logs/verifier/reward.txt").write_text("0.0")
    print("0.0")
    raise SystemExit(0)

ref = load_workbook(ref_file, data_only=True).active
agent = load_workbook(agent_file, data_only=True).active

ref_eigs = [
    complex(ref.cell(r, 2).value, ref.cell(r, 3).value)
    for r in range(2, 5)
]

agent_eigs = [
    complex(agent.cell(r, 2).value, agent.cell(r, 3).value)
    for r in range(2, 5)
]

passed = 0.0

matched = [False] * len(agent_eigs)

for ref_eig in ref_eigs:
    for i, agent_eig in enumerate(agent_eigs):
        if matched[i]:
            continue

        if (
            abs(ref_eig.real - agent_eig.real) <= tol
        ):
            passed += 1.0 / 6.0
            if (abs(ref_eig.imag - agent_eig.imag) <= tol):
                matched[i] = True
                passed += 1.0 / 6.0
                break

for c in range(1, 4):
    a = ref.cell(7, c).value
    b = agent.cell(7, c).value

    if abs(a - b) <= tol:
        passed += 1.0 / 3.0

score = round(passed / 2.0, 4)

Path("/logs/verifier/reward.txt").write_text(str(score))
print(score)