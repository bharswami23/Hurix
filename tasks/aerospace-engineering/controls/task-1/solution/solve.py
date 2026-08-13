import mpmath as mp
from pathlib import Path
from openpyxl import load_workbook, Workbook

mp.mp.dps = 80
alpha = mp.mpf("0.95")
xlsx = Path("/workspace/data/Parameters.xlsx")

wb = load_workbook(xlsx, data_only=True)
env = wb["Environmental Parameters"]
uav = wb["UAV parameters"]

def param(ws, name):
    for row in ws.iter_rows(values_only=True):
        if len(row) >= 3 and row[1] == name:
            return row[2]
    raise KeyError(name)

beta = mp.mpf(str(param(env, "Linear Wind Shear (β)")))
g = mp.mpf(str(param(env, "Acceleration due to gravity (g)")))
rho = mp.mpf(str(param(env, "Density of air (ρ)")))

m = mp.mpf(str(param(uav, "Mass (m)")))
S = mp.mpf(str(param(uav, "Reference Area (S)")))
CL = mp.mpf(str(param(uav, "Lift Coefficient (CL)")))
cd0 = mp.mpf(str(param(uav, "Drag Coefficient (CD0)")))
cd1 = mp.mpf(str(param(uav, "Drag Coefficient (CD1)")))
cd2 = mp.mpf(str(param(uav, "Drag Coefficient (CD2)")))
T = mp.mpf(str(param(uav, "Thrust (T)")))

CD = cd0 + cd1 * CL + cd2 * CL**2
psi = mp.pi / 2

def E(V, gam, mu):
    e1 = (
        T - mp.mpf("0.5") * rho * S * V**2 * CD
        - m * g * mp.sin(gam)
        - m * beta * V * mp.sin(gam) * mp.sin(psi) * mp.cos(gam)
    ) / m

    e2 = (
        mp.mpf("0.5") * rho * S * CL * V**2 * mp.cos(mu)
        - m * g * mp.cos(gam)
        + m * beta * V * mp.sin(gam)**2 * mp.sin(psi)
    ) / (m * V)

    e3 = (
        mp.mpf("0.5") * rho * S * V**2 * CL * mp.sin(mu)
        - m * beta * V * mp.sin(gam) * mp.cos(psi)
    ) / (m * V * mp.cos(gam))

    return e1, e2, e3

V0, gam0, mu0 = mp.findroot(E, (mp.mpf("20"), mp.mpf("-0.05"), mp.mpf("0")), tol=mp.mpf("1e-50"), maxsteps=200)

if not (V0 > 0 and -mp.pi/2 <= gam0 <= mp.pi/2 and -mp.pi/2 <= mu0 <= mp.pi/2):
    raise RuntimeError("Steady solution violates the required bounds.")

def H(a, b, z):
    return mp.hyper([a], b, z)

def HR(a, b, z):
    return mp.hyper([a], b, z) / mp.gamma(b[0]) / mp.gamma(b[1])

def evaluate_A(V, gam, psi, mu, CL, CD, T):
    c1 = mp.mpf("0.0513608432635838663138372339745")
    c2 = mp.mpf("0.0256804216317919331569186169872")
    c3 = mp.mpf("0.978301776449215787501145769056")
    c4 = mp.mpf("1.02721686527167643809832497936")
    c5 = mp.mpf("1.71207756593237947484453798097")
    c6 = mp.mpf("1.25172558362154684452605124534")
    c7 = mp.mpf("0.513608432635838219049162489682")
    c8 = mp.mpf("0.0487928011004046743859241530572")
    c9 = mp.mpf("3.42415513186475894968907596194")
    c10 = mp.mpf("1.95660355289843068682387183799")
    c11 = mp.mpf("0.618136090677307348251190433075")

    A11 = (
        c1*T - c3*CD*S*V**2*rho
        + mp.sin(gam)*(-c1*g*m - c4*m*V*beta*mp.cos(gam)*mp.sin(psi))
    ) / (m*V**alpha)

    A12 = (
        c1*T - c2*CD*S*V**2*rho
        - c5*g*m*gam*HR(1, [mp.mpf("0.525000000000000022204460492503"), mp.mpf("1.02499999999999991118215802999")], -gam**2/4)
        - c4*m*V*beta*gam*H(1, [mp.mpf("1.52499999999999991118215802999"), mp.mpf("1.02499999999999991118215802999")], -gam**2)*mp.sin(psi)
        + c6*m*V*beta*gam**3*H(2, [mp.mpf("2.52499999999999991118215802999"), mp.mpf("2.02499999999999991118215802999")], -gam**2)*mp.sin(psi)
    ) / (m*gam**alpha)

    A13 = (
        c1*T - c2*CD*S*V**2*rho - c1*g*m*mp.sin(gam)
        - c5*m*V*beta*psi*mp.cos(gam)
        * HR(1, [mp.mpf("0.525000000000000022204460492503"), mp.mpf("1.02499999999999991118215802999")], -psi**2/4)
        * mp.sin(gam)
    ) / (m*psi**alpha)

    A21 = (
        c7*CL*S*V**mp.mpf("0.0500000000000000444089209850063")*rho*mp.cos(mu)/m
        + c8*g*mp.cos(gam)*(mp.mpf("18.8679977474509499302257609088")+mp.log(V))/V**mp.mpf("1.94999999999999995559107901499")
        + c1*beta*mp.sin(gam)**2*mp.sin(psi)/V**alpha
    )

    A22 = (
        c2*CL*S*V*rho*mp.cos(mu)/(m*gam**alpha)
        - c9*g*HR(1, [mp.mpf("0.0250000000000000222044604925031"), mp.mpf("0.525000000000000022204460492503")], -gam**2/4)/(V*gam**alpha)
        + beta*(
            c10*gam**mp.mpf("1.04999999999999982236431605997")*H(1, [mp.mpf("2.02499999999999991118215802999"), mp.mpf("1.52499999999999991118215802999")], -gam**2)
            - c11*gam**mp.mpf("3.04999999999999982236431605997")*H(2, [mp.mpf("3.02499999999999991118215802999"), mp.mpf("2.52499999999999991118215802999")], -gam**2)
        )*mp.sin(psi)
    )

    A23 = (
        (-c1*g*mp.cos(gam) + c2*CL*S*V**2*rho*mp.cos(mu)/m)/(V*psi**alpha)
        + c5*beta*psi**mp.mpf("0.0500000000000000444089209850063")
        * HR(1, [mp.mpf("0.525000000000000022204460492503"), mp.mpf("1.02499999999999991118215802999")], -psi**2/4)
        * mp.sin(gam)**2
    )

    A31 = (
        c7*CL*S*V*rho/mp.cos(gam)*mp.sin(mu)
        - c1*m*beta*mp.cos(psi)*mp.tan(gam)
    ) / (m*V**alpha)

    A32 = mp.mpc(0)

    A33 = (
        c2*CL*S*V*rho/mp.cos(gam)*mp.sin(mu)
        - c9*m*beta*HR(1, [mp.mpf("0.0250000000000000222044604925031"), mp.mpf("0.525000000000000022204460492503")], -psi**2/4)*mp.tan(gam)
    ) / (m*psi**alpha)

    return mp.matrix([[A11, A12, A13], [A21, A22, A23], [A31, A32, A33]])

def evaluate_B(V, gam, psi, mu, CL, cd0, cd1, cd2, T):
    c1 = mp.mpf("0.0513608432635838663138372339745")
    c2 = mp.mpf("0.0256804216317919331569186169872")
    c4 = mp.mpf("1.02721686527167643809832497936")
    c5 = mp.mpf("1.71207756593237969688914290600")
    c7 = mp.mpf("0.513608432635838219049162489682")
    c12 = mp.mpf("0.856038782966189737422268990485")

    B11 = (
        c1*T
        + (
            -c2*cd0
            -c7*cd1*CL
            -mp.mpf("0.978301776449215787501145769056")*cd2*CL**2
        ) * S * V**2 * rho
        + mp.sin(gam) * (
            -c1*g*m
            -c1*m*V*beta*mp.cos(gam)*mp.sin(psi)
        )
    ) / (CL**alpha * m)

    B13 = (
        c4*T
        + (
            -c2*cd0
            -c2*cd1*CL
            -c2*cd2*CL**2
        ) * S * V**2 * rho
        + mp.sin(gam) * (
            -c1*g*m
            -c1*m*V*beta*mp.cos(gam)*mp.sin(psi)
        )
    ) / (m*T**alpha) if T != 0 else mp.inf

    B21 = (
        -c1*g*m*mp.cos(gam)
        + V * (
            c7*CL*S*V*rho*mp.cos(mu)
            + c1*m*beta*mp.sin(gam)**2*mp.sin(psi)
        )
    ) / (CL**alpha * m * V)

    B22 = (
        -c1*g*m*mp.cos(gam)
        + c5*CL*S*V**2*rho
        * HR(
            1,
            [
                mp.mpf("0.0250000000000000222044604925031"),
                mp.mpf("0.525000000000000022204460492503")
            ],
            -mu**2
        )
        + c1*m*V*beta*mp.sin(gam)**2*mp.sin(psi)
    ) / (m*V*mu**alpha)

    B31 = (
        c7*CL*S*V*rho/mp.cos(gam)*mp.sin(mu)
        - c1*m*beta*mp.cos(psi)*mp.tan(gam)
    ) / (CL**alpha * m)

    B32 = (
        CL*S*V*mu**mp.mpf("0.0500000000000000444089209850063")*rho
        * HR(
            1,
            [
                mp.mpf("0.525000000000000022204460492503"),
                mp.mpf("1.02499999999999991118215802999")
            ],
            -mu**2/4
        ) * (1/mp.cos(gam))
        - c1*beta*mp.cos(psi)*mp.tan(gam)/mu**alpha
    ) / (m*c12)

    return mp.matrix([
        [B11, mp.mpf(0), B13],
        [B21, B22, mp.mpf(0)],
        [B31, B32, mp.mpf(0)]
    ])

A = evaluate_A(V0, gam0, psi, mu0, CL, CD, T)
B = evaluate_B(V0, gam0, psi, mu0, CL, cd0, cd1, cd2, T)

def charpoly_coeffs(M):
    tr = M[0,0] + M[1,1] + M[2,2]
    c1p = (
        M[0,0]*M[1,1] + M[0,0]*M[2,2] + M[1,1]*M[2,2]
        - M[0,1]*M[1,0] - M[0,2]*M[2,0] - M[1,2]*M[2,1]
    )
    det = (
        M[0,0]*(M[1,1]*M[2,2]-M[1,2]*M[2,1])
        - M[0,1]*(M[1,0]*M[2,2]-M[1,2]*M[2,0])
        + M[0,2]*(M[1,0]*M[2,1]-M[1,1]*M[2,0])
    )
    return [mp.mpc(1), -tr, c1p, -det]

b = mp.matrix([B[0,0], B[1,0], B[2,0]])

def eigs(k):
    K = mp.matrix([[mp.mpf(0), k, mp.mpf(0)]])
    M = A - b*K
    return mp.polyroots(charpoly_coeffs(M), maxsteps=10000, error=False)

def most_stable(k):
    roots = eigs(k)
    return min(mp.re(z) for z in roots)

klow = mp.mpf("0")
kmid = mp.mpf("1")
khigh = mp.mpf("2")
khighprev = khigh

for _ in range(300):
    shigh = most_stable(khigh)
    if abs(shigh + 1) < mp.mpf("1e-30"):
        break
    if shigh < -1:
        khighprev = khigh
        khigh = kmid
        kmid = (klow + khigh) / 2
    else:
        khigh = khighprev
        klow = kmid
        kmid = (klow + khigh) / 2

lam = eigs(khigh)

out = Path("/logs/agent")
out.mkdir(parents=True, exist_ok=True)

wb_out = Workbook()
ws = wb_out.active
ws.title = "Results"
ws.append(["Mode number", "Real Part of Eigenvalue", "Imaginary Part of Eigenvalue"])

for i, z in enumerate(lam, 1):
    ws.append([i, round(float(mp.re(z)), 6), round(float(mp.im(z)), 6)])

ws.append(["", "", ""])
ws["A6"] = "Feedback Matrix:"
ws["A7"] = round(float(0), 6)
ws["B7"] = round(float(khigh), 6)
ws["C7"] = round(float(0), 6)

wb_out.save(out / "results.xlsx")

for i, z in enumerate(lam, 1):
    print(i, mp.nstr(mp.re(z), 20), mp.nstr(mp.im(z), 20))
print("k1", mp.nstr(khigh, 20))
print("Output", out / "results.xlsx")